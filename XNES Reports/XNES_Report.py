import os
import pyodbc
import pandas as pd
import openpyxl as pyxl
from openpyxl.worksheet.worksheet import Worksheet
from pyodbc import Connection, Cursor
from datetime import datetime, timedelta
from dotenv import load_dotenv
from typing import Tuple, List, Optional

# Load environment variables
load_dotenv()

# Constants
RUN_DATE: datetime = datetime.today()
# Get the last Friday
QUERY_DATE: datetime = RUN_DATE - timedelta(days=(1 + (RUN_DATE.weekday() + 2) % 7))

CONN_STRING: str = os.getenv("CONN_STRING", "")
EQUITY_ID: str = os.getenv("EQUITY_ID", "")

# Stored procedures
EXPOSURES_STORED_PROC: str = os.getenv("EXPOSURES_STORED_PROC", "")
MARGIN_STORED_PROC: str = os.getenv("MARGIN_STORED_PROC", "")
VAR_STORED_PROC: str = os.getenv("VAR_STORED_PROC", "")
POS_CHANGES_STORED_PROC: str = os.getenv("POS_CHANGES_STORED_PROC", "")


def fetch_data_from_db(conn: Connection, stored_proc: str, params: Tuple,
                       column_names: Optional[List[str]] = None) -> pd.DataFrame:
    """Fetch data from the database using a stored procedure and return it as a DataFrame."""
    curr: Cursor = conn.cursor()
    curr.execute(stored_proc, params)
    rows = curr.fetchall()

    # If column_names are provided, use them; otherwise, extract from cursor
    if column_names:
        col_names = column_names
    else:
        col_names = [column[0] for column in curr.description]

    curr.close()
    return pd.DataFrame.from_records(rows, columns=col_names)


def get_exposures(conn: Connection, equity_id: str, query_date: datetime) -> pd.DataFrame:
    """Fetch exposures data and return as a DataFrame."""
    params: Tuple[str, float, str] = (equity_id, 0.0, query_date.strftime("%Y-%m-%d"))
    return fetch_data_from_db(conn, EXPOSURES_STORED_PROC, params)


def get_margin(conn: Connection, query_date: datetime, equity_id: str) -> pd.DataFrame:
    """Fetch margin data and return as a DataFrame."""
    params: Tuple[str, str] = (query_date.strftime("%Y-%m-%d"), equity_id)
    return fetch_data_from_db(conn, MARGIN_STORED_PROC, params,
                              column_names=["Date", "VarType", "Margin", "System", "Account", "Sector", "Total M2E"])


def get_var(conn: Connection, query_date: datetime, equity_id: str) -> pd.DataFrame:
    """Fetch VaR data and return as a DataFrame."""
    params: Tuple[str, str] = (query_date.strftime("%Y-%m-%d"), equity_id)
    return fetch_data_from_db(conn, VAR_STORED_PROC, params, column_names=["Sector", "VaR", "TotalVaR"])


def get_position_changes(conn: Connection, start_date: datetime, end_date: datetime, equity_id: str) -> pd.DataFrame:
    """Fetch position changes data and return as a DataFrame."""
    params: Tuple[str, str, str] = (start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d"), equity_id)
    return fetch_data_from_db(conn, POS_CHANGES_STORED_PROC, params,
                              column_names=["Code", "Sector", "Name", "qty", "dsc", "date", "PnL"])


def populate_exposures(ws: Worksheet, exposures_df: pd.DataFrame) -> None:
    """Populate the exposures section in the Excel sheet."""
    grouped_exposures_df = exposures_df[['סקטור', 'חשיפה', 'חשיפה נטו']].groupby('סקטור').sum()
    grouped_exposures_df.sort_index(ascending=False, inplace=True)
    ws['I3'] = exposures_df['חשיפה'].sum()
    ws['I4'] = exposures_df['חשיפה נטו'].sum()
    ws['I5'] = exposures_df['חשיפה נטו'].where(exposures_df['חשיפה נטו'] > 0, 0).sum()
    ws['I6'] = exposures_df['חשיפה נטו'].where(exposures_df['חשיפה נטו'] < 0, 0).sum() * -1

    grouped_exposures_df['ofTotal'] = grouped_exposures_df['חשיפה'] / grouped_exposures_df['חשיפה'].sum()

    for i, (index, row) in enumerate(grouped_exposures_df.iterrows()):
        ws[f'J{i + 9}'] = index
        ws[f'I{i + 9}'] = row['חשיפה נטו']
        ws[f'H{i + 9}'] = row['חשיפה']
        ws[f'G{i + 9}'] = row['ofTotal']

    for i, (index, row) in enumerate(exposures_df.iterrows()):
        if i == 5:
            break
        ws[f'D{i + 3}'] = row['שוק']
        ws[f'C{i + 3}'] = row['חשיפה']
        ws[f'B{i + 3}'] = row['צד']


def populate_margin(ws: Worksheet, margin_df: pd.DataFrame) -> None:
    """Populate the margin section in the Excel sheet."""
    margin_df = margin_df.loc[~margin_df['Sector'].isin(['סה"כ', 'ריביות'])]
    margin_df.sort_values(by='Sector', ascending=False, inplace=True, ignore_index=True)
    ws["I14"] = margin_df['Margin'].sum()
    for i, (index, row) in enumerate(margin_df.iterrows()):
        ws[f'J{i + 15}'] = row['Sector']
        ws[f'I{i + 15}'] = row['Margin']


def populate_var(ws: Worksheet, var_df: pd.DataFrame) -> None:
    """Populate the VaR section in the Excel sheet."""
    total_var = var_df['TotalVaR'].sum() / 10000
    var_df = var_df.loc[~var_df['Sector'].isin(['סה"כ', 'ריביות'])]
    var_df['VaR'] = var_df['VaR'] * total_var
    var_df.sort_values(by='Sector', ascending=False, inplace=True, ignore_index=True)
    ws["F14"] = var_df['VaR'].sum()
    for i, (index, row) in enumerate(var_df.iterrows()):
        ws[f'G{i + 15}'] = row['Sector']
        ws[f'F{i + 15}'] = row['VaR']


def populate_position_changes(ws: Worksheet, pos_changes_df: pd.DataFrame) -> None:
    """Populate the position changes section in the Excel sheet."""
    for i, (index, row) in enumerate(pos_changes_df.iterrows()):
        ws[f'E{i + 23}'] = row['Code']
        ws[f'F{i + 23}'] = row['Sector']
        ws[f'G{i + 23}'] = row['Name']
        ws[f'H{i + 23}'] = row['qty']
        ws[f'I{i + 23}'] = row['dsc']
        ws[f'J{i + 23}'] = row['date']


if __name__ == "__main__":
    # Connect to the database
    conn: Connection = pyodbc.connect(CONN_STRING)

    # Load the Excel template
    wb: pyxl.Workbook = pyxl.load_workbook("TEMPLATE.xlsx")
    ws: Worksheet = wb.active
    ws.title = "SnapShot"

    # Populate date in the Excel
    ws['J2'] = QUERY_DATE.strftime("%Y-%m-%d")

    # Fetch and populate data
    exposures_df = get_exposures(conn, EQUITY_ID, QUERY_DATE)
    populate_exposures(ws, exposures_df)

    margin_df = get_margin(conn, QUERY_DATE, EQUITY_ID)
    populate_margin(ws, margin_df)

    var_df = get_var(conn, QUERY_DATE, EQUITY_ID)
    populate_var(ws, var_df)

    start_date = QUERY_DATE - timedelta(days=7)
    pos_changes_df = get_position_changes(conn, start_date, QUERY_DATE, EQUITY_ID)
    populate_position_changes(ws, pos_changes_df)

    # Close the database connection
    conn.close()

    # Save the workbook
    report_filename = f"XNES-COMM-WEEKLY_{QUERY_DATE.strftime('%Y%m%d')}.xlsx"
    wb.save(report_filename)
    print(f"Report generated successfully: {report_filename}")
